from openpyxl import load_workbook;
from openpyxl.chart import BarChart, Reference;

#leitura de pasta de trabalho 
wb = load_workbook("Data/pivot_table.xlsx");

sheet = wb["Relatorio"];

#Referencias de linhas e colunas
min_colum = wb.active.min_column;
max_colum = wb.active.max_column;

min_row = wb.active.min_row;
max_row = wb.active.max_row;

print(min_colum, max_colum);
print(min_row, max_row);

#Dados e categorias em um grafico

BarChart = BarChart();

data = Reference(
    sheet,
    max_col = max_colum,
    min_col = min_colum+1,
    max_row = min_row,
    min_row= min_row,
);

categories = Reference(
    sheet,
    max_col = min_colum,
    min_col = min_colum,
    max_row = min_row + 1,
    min_row= min_row,  
);

BarChart.add_data(data, titles_from_data= True);
BarChart.set_categories(categories);

#Criando grafico
sheet.add_chart(BarChart, "B10");
BarChart.title = "Vendas por Fabricante";
BarChart.style = 2;

#Salvando o workbook
wb.save("data/barchart.xlsx");

