from openpyxl import load_workbook;
from openpyxl.utils import get_column_letter;

#leitura de pasta de trabalho 
wb = load_workbook("Data/barchart.xlsx");

sheet = wb["Relatorio"];

#Referencias de linhas e colunas
min_colum = wb.active.min_column;
max_colum = wb.active.max_column;
min_row = wb.active.min_row;
max_row = wb.active.max_row;

#Usando Formulas
# sheet["B6"] = "=SUM(B2:B5)";
# sheet["B6"].style = "Currency";
for i in range(min_colum + 1, max_colum + 1):
    letter = get_column_letter(i);
    # print(letter);
    sheet[f"{letter}{max_row + 1}"] = f"=SUM({letter}{min_row + 1}:{letter}{max_row})";
    sheet[f"{letter}{max_row + 1}"].style = "Currency";

wb.save("teste.xlsx");